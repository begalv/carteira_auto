"""Exportadores para planilhas Excel."""

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from carteira_auto.config import constants, settings
from carteira_auto.core.models import Portfolio
from carteira_auto.utils import get_logger
from carteira_auto.utils.decorators import log_execution, timer

logger = get_logger(__name__)


class ExcelExporter:
    """Exportador genérico que copia uma planilha e aplica modificações.

    Preserva formatação e fórmulas da planilha original.
    Subclasses implementam a lógica de atualização específica.

    Usage:
        exporter = ExcelExporter(source, output)
        with exporter:
            ws = exporter.get_sheet("Aba1")
            ws.cell(row=2, column=1, value="dado")
        # Salva automaticamente ao sair do context manager

        # Ou manualmente:
        exporter.open()
        ws = exporter.get_sheet("Aba1")
        exporter.save()
        exporter.close()
    """

    def __init__(self, source_path: Path, output_path: Path):
        self.source_path = source_path
        self.output_path = output_path
        self._wb: Workbook | None = None

    def open(self) -> "ExcelExporter":
        """Copia a planilha original e abre a cópia para edição."""
        if not self.source_path.exists():
            raise FileNotFoundError(
                f"Planilha original não encontrada: {self.source_path}"
            )
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(self.source_path, self.output_path)
        self._wb = load_workbook(self.output_path)
        logger.debug(f"Planilha aberta para exportação: {self.output_path}")
        return self

    def save(self) -> None:
        """Salva as modificações no arquivo de saída."""
        self._ensure_open()
        self._wb.save(self.output_path)
        logger.debug(f"Planilha salva: {self.output_path}")

    def close(self) -> None:
        """Fecha o workbook."""
        if self._wb is not None:
            self._wb.close()
            self._wb = None

    def __enter__(self) -> "ExcelExporter":
        return self.open()

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        if exc_type is None:
            self.save()
        self.close()

    @property
    def sheet_names(self) -> list[str]:
        """Retorna os nomes das abas disponíveis."""
        self._ensure_open()
        return self._wb.sheetnames

    def get_sheet(self, sheet_name: str, required: bool = True):
        """Obtém uma aba do workbook.

        Args:
            sheet_name: Nome da aba.
            required: Se True, levanta erro quando não existe.
                      Se False, retorna None.
        """
        self._ensure_open()
        if sheet_name not in self._wb.sheetnames:
            if required:
                raise ValueError(f"Aba '{sheet_name}' não encontrada")
            logger.warning(f"Aba '{sheet_name}' não encontrada, ignorando")
            return None
        return self._wb[sheet_name]

    def _ensure_open(self) -> None:
        """Garante que o workbook está aberto."""
        if self._wb is None:
            raise RuntimeError(
                "Planilha não aberta. Use open() ou o context manager 'with'."
            )


class PortfolioPriceExporter(ExcelExporter):
    """Exporta preços atuais para a planilha da carteira.

    Copia a planilha original e atualiza apenas a coluna 'Preço Atual',
    preservando toda a formatação e fórmulas existentes.

    Usage:
        exporter = PortfolioPriceExporter()  # usa paths padrão de settings
        output = exporter.export_prices(portfolio)

        exporter = PortfolioPriceExporter(
            source_path=Path("minha/planilha.xlsx"),
            output_path=Path("saida/atualizada.xlsx"),
        )
        output = exporter.export_prices(portfolio)
    """

    _PRECO_ATUAL_COL = constants.CARTEIRA_PRECO_ATUAL_COL
    _TICKER_COL = constants.CARTEIRA_TICKER_COL

    def __init__(
        self,
        source_path: Path | None = None,
        output_path: Path | None = None,
    ):
        src = source_path or settings.paths.PORTFOLIO_FILE
        out = output_path or settings.paths.get_portfolio_output_path()
        super().__init__(src, out)

    @log_execution
    @timer
    def export_prices(self, portfolio: Portfolio) -> Path:
        """Atualiza a coluna 'Preço Atual' com os preços do portfolio.

        Returns:
            Path do arquivo exportado.
        """
        price_map = {
            a.ticker: a.preco_atual
            for a in portfolio.assets
            if a.preco_atual is not None
        }

        if not price_map:
            logger.warning("Nenhum preço atualizado no portfolio")
            return self.output_path

        with self:
            ws = self.get_sheet(constants.CARTEIRA_SHEET_NAMES["carteira"])
            updated = 0
            for row in range(2, ws.max_row + 1):
                ticker = ws.cell(row=row, column=self._TICKER_COL).value
                if ticker in price_map:
                    ws.cell(
                        row=row,
                        column=self._PRECO_ATUAL_COL,
                        value=round(price_map[ticker], 2),
                    )
                    updated += 1

        logger.info(
            f"Preços exportados: {updated}/{len(price_map)} ativos "
            f"atualizados em {self.output_path}"
        )
        return self.output_path
